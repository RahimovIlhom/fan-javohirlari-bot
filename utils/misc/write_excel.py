from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


async def write_data_excel(columns, data):
    workbook = Workbook()
    sheet = workbook.active

    # Add styles
    header_style = Font(bold=True)
    align_center = Alignment(horizontal='center', vertical='center')

    # Write header with styles
    for col_num, col_value in enumerate(columns, 1):
        cell = sheet.cell(row=1, column=col_num, value=col_value)
        cell.font = header_style
        cell.alignment = align_center

    # Write data with styles
    for row_num, row_values in enumerate(data, 2):
        for col_num, col_value in enumerate(row_values, 1):
            cell = sheet.cell(row=row_num, column=col_num, value=col_value)
            cell.alignment = align_center

    workbook.save('data/users/data.xlsx')
